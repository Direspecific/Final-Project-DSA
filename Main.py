import datetime
import tkinter 
import customtkinter
import math
import pandas as pd
import os
from CTkMessagebox import CTkMessagebox
from PIL import Image
import openpyxl
import collections as clc

def side_panel_switch():
    if (math.floor(title_frame.winfo_x()) == -75):
        protract()
        return
    if (math.floor(title_frame.winfo_x()) == 75):   
        retract()
        return
    
def protract():
    title_frame.place(x=title_frame.winfo_x()+15)
    home_frame.place(x=(title_frame.winfo_x())+30) 
    lithit_main_frame.place(x=(title_frame.winfo_x())+35)
    library_main_frame.place(x=(title_frame.winfo_x())+75) 
    add_books_main_frame.place(x=(title_frame.winfo_x()+60)) 
    print(home_frame.winfo_x())
    search_advanceSearch_frame.place(x=(title_frame.winfo_x())+15)
    search_library_frame_inner.place(x=(title_frame.winfo_x()+615))

    if math.floor(title_frame.winfo_x()) == 60:
        Main_app.after_cancel(protract)
    else:
        Main_app.after(10,protract)
    
def retract():
    title_frame.place(x=title_frame.winfo_x()-15)
    home_frame.place(x=(title_frame.winfo_x())+45) 
    lithit_main_frame.place(x=(title_frame.winfo_x())+60)
    library_main_frame.place(x=(title_frame.winfo_x())+45) 
    add_books_main_frame.place(x=(title_frame.winfo_x()+80)) 
    print(home_frame.winfo_x())
    search_advanceSearch_frame.place(x=title_frame.winfo_x()-15)
    search_library_frame_inner.place(x=(title_frame.winfo_x()+585))

    if  math.floor(title_frame.winfo_x()) == -60:
        Main_app.after_cancel(retract)
        
    else:
        Main_app.after(10,retract)


def show_main_frame_content(text):

    if text == "Home":
        library_frame.pack_forget()
        search_frame.pack_forget()
        lithit_frame.pack_forget()
        add_books_frame.pack_forget()
        home_frame.pack(fill='both', expand=1)
        
        button_home.configure(state="disabled")
        button_lithit.configure(state="normal")
        button_library.configure(state="normal")
        button_search.configure(state="normal")
        button_addbooks.configure(state="normal")
        
        if button_home.cget("state")=="disabled":
            button_home_img_clicked = customtkinter.CTkImage((Image.open("icons\homeclicked.png")),size=(45,45))
            button_home.configure(image=button_home_img_clicked)
            button_lithit_img = customtkinter.CTkImage((Image.open("icons\lithitsunclicked.png")),size=(35,45))
            button_lithit.configure(image=button_lithit_img)
            button_library_img = customtkinter.CTkImage((Image.open("icons\libraryunclicked.png")),size=(45,45))
            button_library.configure(image=button_library_img)
            button_search_img = customtkinter.CTkImage((Image.open("icons\searchunclicked.png")),size=(45,45))
            button_search.configure(image=button_search_img)
            button_addbooks_img_clicked = customtkinter.CTkImage((Image.open("icons\disclickedaddbooks.png")),size=(45,45))
            button_addbooks.configure(image=button_addbooks_img_clicked)
        else:
            button_home_img = customtkinter.CTkImage((Image.open("icons\homeunclicked.png")),size=(45,45))
            button_home.configure(image=button_home_img)

        
    if text == "lithit":
        library_frame.pack_forget()
        search_frame.pack_forget()
        home_frame.pack_forget()
        add_books_frame.pack_forget()
        lithit_frame.pack(fill='both', expand=1)

        
        button_home.configure(state="normal")
        button_lithit.configure(state="disabled")
        button_library.configure(state="normal")
        button_search.configure(state="normal")
        button_addbooks.configure(state="normal")
        
        if button_lithit.cget("state")=="disabled":
            button_home_img = customtkinter.CTkImage((Image.open("icons\homeunclicked.png")),size=(45,45))
            button_home.configure(image=button_home_img)
            button_lithit_img_clicked = customtkinter.CTkImage((Image.open("icons\lithitsclicked.png")),size=(35,45))
            button_lithit.configure(image=button_lithit_img_clicked)
            button_library_img = customtkinter.CTkImage((Image.open("icons\libraryunclicked.png")),size=(45,45))
            button_library.configure(image=button_library_img)
            button_search_img = customtkinter.CTkImage((Image.open("icons\searchunclicked.png")),size=(45,45))
            button_search.configure(image=button_search_img)
            button_addbooks_img_clicked = customtkinter.CTkImage((Image.open("icons\disclickedaddbooks.png")),size=(45,45))
            button_addbooks.configure(image=button_addbooks_img_clicked)
        else:
            button_lithit_img = customtkinter.CTkImage((Image.open("icons\lithitsunclicked.png")),size=(45,45))
            button_lithit.configure(image=button_lithit_img)
        
    if text == "Search":
        library_frame.pack_forget()
        lithit_frame.pack_forget()
        home_frame.pack_forget()
        add_books_frame.pack_forget()
        search_frame.pack(fill='both', expand=1)
        
        button_home.configure(state="normal")
        button_lithit.configure(state="normal")
        button_search.configure(state="disabled")
        button_library.configure(state="normal")
        button_addbooks.configure(state="normal")
        
        if button_search.cget("state")=="disabled":
            button_home_img = customtkinter.CTkImage((Image.open("icons\homeunclicked.png")),size=(45,45))
            button_home.configure(image=button_home_img)
            button_lithit_img = customtkinter.CTkImage((Image.open("icons\lithitsunclicked.png")),size=(35,45))
            button_lithit.configure(image=button_lithit_img)
            button_search_img_clicked = customtkinter.CTkImage((Image.open("icons\searchclicked.png")),size=(45,45))
            button_search.configure(image=button_search_img_clicked)
            button_library_img = customtkinter.CTkImage((Image.open("icons\libraryunclicked.png")),size=(45,45))
            button_library.configure(image=button_library_img)
            button_addbooks_img_clicked = customtkinter.CTkImage((Image.open("icons\disclickedaddbooks.png")),size=(45,45))
            button_addbooks.configure(image=button_addbooks_img_clicked)

        else:
            button_search_img = customtkinter.CTkImage((Image.open("icons\\searchunclicked.png")),size=(45,45))
            button_search.configure(image=button_search_img)

    if text == "Library":
        search_frame.pack_forget()
        lithit_frame.pack_forget()
        home_frame.pack_forget()
        add_books_frame.pack_forget()
        library_frame.pack(fill='both', expand=1)

        button_home.configure(state="normal")
        button_lithit.configure(state="normal")
        button_search.configure(state="normal")
        button_library.configure(state="disabled")
        button_addbooks.configure(state="normal")
        
        if button_library.cget("state")=="disabled":
            button_home_img = customtkinter.CTkImage((Image.open("icons\homeunclicked.png")),size=(45,45))
            button_home.configure(image=button_home_img)
            button_lithit_img = customtkinter.CTkImage((Image.open("icons\\lithitsunclicked.png")),size=(35,45))
            button_lithit.configure(image=button_lithit_img)
            button_search_img_ = customtkinter.CTkImage((Image.open("icons\\searchunclicked.png")),size=(45,45))
            button_search.configure(image=button_search_img_)
            button_library_img_clicked = customtkinter.CTkImage((Image.open("icons\\libraryclicked.png")),size=(45,45))
            button_library.configure(image=button_library_img_clicked)
            button_addbooks_img_clicked = customtkinter.CTkImage((Image.open("icons\\disclickedaddbooks.png")),size=(45,45))
            button_addbooks.configure(image=button_addbooks_img_clicked)

        else:
            button_library_img = customtkinter.CTkImage((Image.open("icons\\libraryunclicked.png")),size=(45,45))
            button_library.configure(image=button_library_img)

    if text == "Add":
        search_frame.pack_forget()
        lithit_frame.pack_forget()
        home_frame.pack_forget()
        library_frame.pack_forget()
        add_books_frame.pack(fill='both', expand=1)
        
        button_home.configure(state="normal")
        button_lithit.configure(state="normal")
        button_search.configure(state="normal")
        button_library.configure(state="normal")
        button_addbooks.configure(state="disabled")
        
        if button_addbooks.cget("state")=="disabled":
            button_home_img = customtkinter.CTkImage((Image.open("icons\\homeunclicked.png")),size=(45,45))
            button_home.configure(image=button_home_img)
            button_lithit_img = customtkinter.CTkImage((Image.open("icons\\lithitsunclicked.png")),size=(35,45))
            button_lithit.configure(image=button_lithit_img)
            button_search_img_ = customtkinter.CTkImage((Image.open("icons\\searchunclicked.png")),size=(45,45))
            button_search.configure(image=button_search_img_)
            button_library_img = customtkinter.CTkImage((Image.open("icons\\libraryunclicked.png")),size=(45,45))
            button_library.configure(image=button_library_img)
            button_addbooks_img_clicked = customtkinter.CTkImage((Image.open("icons\\clickedaddbooks.png")),size=(45,45))
            button_addbooks.configure(image=button_addbooks_img_clicked)

        else:
            button_addbooks_img = customtkinter.CTkImage((Image.open("icons\\disclickedaddbooks.png")),size=(45,45))
            button_addbooks.configure(image=button_addbooks_img)
    # else:
    #     search_frame.pack_forget()
    #     lithit_frame.pack_forget()
    #     home_frame.pack_forget()
    #     add_books_frame.pack_forget()
    #     library_frame.pack_forget()
    #     log_in_frame.pack(fill='both', expand=1)    

def log_in_clock():
    
    rtc = datetime.datetime.now()
    if (rtc.hour < 12):
        if (rtc.hour == 12):
            hr_min_time = "{:02d}:{:02d} AM".format(rtc.hour+12,rtc.minute)
        else:
            hr_min_time = "{:02d}:{:02d} AM".format(rtc.hour,rtc.minute)
    else:
        hr_min_time = "{:02d}:{:02d} PM".format(rtc.hour-12,rtc.minute)
    login_clock_label.configure(text=hr_min_time)
    Main_app.after(1000,log_in_clock)

def logout():
    home_frame.pack_forget()
    library_frame.pack_forget()
    search_frame.pack_forget()
    lithit_frame.pack_forget()
    add_books_frame.pack_forget()
    log_in_frame.pack(fill='both', expand=1) 

def to_log_in():
    sign_in_frame.pack_forget()
    log_in_frame.pack(fill='both', expand=1)  

def login():
    username = login_username_entry.get()
    password = login_password_entry.get()
    if username in user_credentials and user_credentials[username] == password:
        msg = CTkMessagebox(title="LogIn", message="Login Success!",icon="warning", option_1="Ok")
        response = msg.get()
        if response == "Ok":
            log_in_frame.pack_forget() 
            show_main_frame_content("Home")
            login_username_entry.delete(0,tkinter.END)
            login_password_entry.delete(0,tkinter.END)

    else:
        msg = CTkMessagebox(title="Error Account not found", message="Account Not Found!",icon="warning", option_1="Ok")

def agree():
    value = sign_in_checkbox.get()
    print(value)
    if value == "agree":
        terms_n_conditions = True
        return terms_n_conditions 
    if value == "disagree":
        terms_n_conditions = False
        return terms_n_conditions
    
def to_sign_in():
    log_in_frame.pack_forget()
    sign_in_frame.pack(fill='both', expand=1)    

def sign_in():
    username = sign_in_username_entry.get()
    password = sign_in_password_entry.get()
    email = sign_in_email_entry.get()
    
    if username != "" and password !="" and email != "" and agree() == True :
        if username in user_credentials:
            msg = CTkMessagebox(message="Username is already taken",icon="warning", option_1="Ok")
        else:
            sign_in_username_entry.delete(0,tkinter.END)
            sign_in_password_entry.delete(0,tkinter.END)
            sign_in_email_entry.delete(0,tkinter.END)
            user_credentials[username] = password
            acc_reg_success()
            save_to_excel()
    else:
        msg = CTkMessagebox(message="Please fill  up all the forms",icon="warning", option_1="Ok")
            
def acc_reg_success():
    msg = CTkMessagebox(title="Account Registered", message="Account Registered!",icon="check", option_1="Clock-In")
    response = msg.get()
    
    if response=="Clock-In":
        to_log_in()       

def save_user_credentials():
    dff = pd.DataFrame(user_credentials.items(), columns=['Username', 'Password'])
    dff.to_excel(usercreds_excel_file, index=False)

def add_book():
    title = input("Enter the title of the book: ")
    author = input("Enter the author's name: ")

    # Load existing data from excel
    df = pd.read_excel(books_excel_file)

    # Determine the next available Book ID
    next_book_id = len(df) + 1

    # Create a new DataFrame for the book to be added with the determined Book ID
    new_book = pd.DataFrame({'bookID': [next_book_id], 'Title': [title], 'Author': [author]})

    # Concatenate the new book DataFrame with the existing DataFrame
    df = pd.concat([df, new_book], ignore_index=True)

    # Save the updated data to the excel file, overwriting the previous data
    df.to_excel(books_excel_file, index=False)
    save_user_credentials()
    save_to_excel()
    msg = CTkMessagebox(message="Book Added Succesfully",icon="check", option_1="Ok")
    
def select_book():
    title = input("Enter the title of the book you want to select: ")
    
    # Load existing data from excel
    df = pd.read_excel(books_excel_file)

    # If it's not a valid integer, search by title using str.contains for partial and case-insensitive matching
    found_books = df[df['Title'].str.contains(title, case=False)]
    if found_books.empty:
        print("Book not found in the database.")
    else:
        print("Book Details:")

def delete_book():
    title = input("Enter the title of the book to delete: ")

    # Load existing data from excel
    df = pd.read_excel(books_excel_file)

    try:
        # Create a new DataFrame excluding the book to be deleted
        df = df[df['Title'] != title]

        # Save the updated data to the excel file, overwriting the previous data
        df.to_excel(books_excel_file, index=False)
        save_user_credentials()
        print("Book deleted successfully!")
        save_to_excel()
    except KeyError:
        print("Book is not in the library!")
        
def save_to_excel():
    # Read data from excel into dff DataFrame
    df = pd.read_excel(books_excel_file)

    # Save the data to the Excel file
    df.to_excel(books_excel_file, index=False)
    save_user_credentials()
    print("Data saved to Excel and user credentials saved to excel successfully!")

def update_last_ten_stack():
    last_ten_stack.clear()
    for book_id in list(book_id_linked_list)[-10:]:
        # Find the corresponding row in the Excel file and append it to the stack
        for row in sheet.iter_rows(values_only=True):
            if row[0] == book_id:
                book_info = {
                    'book_id': row[0],
                    'title': row[1],
                    'author': row[2]
                }
                last_ten_stack.append(book_info)
                                              
    display_last_ten_stack_as_buttons()

# Create a function to display the stack as buttons in a scrollable frame in Lit Hits
def display_last_ten_stack_as_buttons():
    
    for book_info in stack:
        book_button = customtkinter.CTkButton(lithit_firstShelf_frame,width=985,hover_color="#B88B68",fg_color="#614D40",height=135, text=f"{book_info['title']} \nby {book_info['author']}",font=("Quando",35), command=lambda info=book_info: display_book_details(info),corner_radius=15)
        book_button.pack(padx=5,pady=5)

# Create a function to display the details of the selected book
def display_book_details(book_info):
    msg = CTkMessagebox(message=f"Book ID: {book_info['book_id']} \n Title: {book_info['title']}\nAuthor: {book_info['author']}",icon="info", option_1="Ok")

# Create a function to display the stack as buttons in a scrollable frame in Library
def display_stack_as_buttons():

    # Create buttons for each book in the stack and add them to the frame
    for book_info in stack:
        book_button = customtkinter.CTkButton(library_bookshelf_frame,width=1200,hover_color="#3D291E",fg_color="#614D40",height=135, text=f"{book_info['title']} \nby {book_info['author']}",font=("Quando",35), command=lambda info=book_info: display_book_details(info),corner_radius=15)
        book_button.pack(padx=5,pady=5)

#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\FOR DATABASE\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#

# File paths
books_excel_file = 'D:\Code\DSA\Final-Project-DSA-main\STACKRead\spreadsheets\\books.xlsx'
usercreds_excel_file = 'D:\Code\DSA\Final-Project-DSA-main\STACKRead\spreadsheets\\usercreds.xlsx'

# Hash table to store username and password
user_credentials = {}

# Create a linked list to store bookId values
book_id_linked_list = clc.deque()

# Create a stack to store the last ten values
last_ten_stack = []

# Load existing user credentials from excel
if os.path.exists(usercreds_excel_file):
    dff = pd.read_excel(usercreds_excel_file)
    for index, row in dff.iterrows():
        user_credentials[row['Username']] = row['Password']
        
# Load existing data from Excel
df = pd.read_excel(books_excel_file)
dff = pd.read_excel(usercreds_excel_file)

# add_book()
# select_book()
# delete_book()
# save_to_excel()

# Create an empty stack to store data
stack = []

# Load the Excel file
workbook = openpyxl.load_workbook(books_excel_file)

# Choose a specific sheet within the Excel file, e.g., the first sheet
sheet = workbook.active

# Use a variable to skip the first row (headers)
skip_first_row = True

# Iterate through the rows in the sheet and add data to the stack
for row in sheet.iter_rows(values_only=True):
    if skip_first_row:
        skip_first_row = False
        continue

    book_info = {
        'book_id': row[0],
        'title': row[1],
        'author': row[2]
    }
    stack.append(book_info)


#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\FOR GUI\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#

customtkinter.set_appearance_mode("light")
customtkinter.set_default_color_theme("blue")

# Main Window
Main_app  = customtkinter.CTk()
Main_app.title("STACKRead")
# Main_app.iconbitmap(Image.open('icons\stackreadicon.png'))

# Get the screen dimensions
screen_width = 1600
screen_height = 900

# Calculate the desired window size as a fraction of the screen dimensions
window_width = 1360  
window_height = 765  

# Set the window size and position it in the center of the screen
x_position = (screen_width - window_width) // 2
y_position = (screen_height - window_height) // 4.5

# Set the window's geometry
Main_app.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")
Main_app.resizable(0,0)

# Main Content Frames
main_content_frame = customtkinter.CTkFrame(Main_app,width=Main_app.winfo_screenwidth(),height=Main_app.winfo_screenheight(),fg_color="#F2E8BD")
main_content_frame.place(x = 50, y = 0)

home_frame = customtkinter.CTkFrame(main_content_frame,width=Main_app.winfo_screenwidth(),height=Main_app.winfo_screenheight(),fg_color="#F2E8BD")
lithit_frame = customtkinter.CTkFrame(main_content_frame,width=Main_app.winfo_screenwidth(),height=Main_app.winfo_screenheight(),fg_color="#F2E8BD")
search_frame = customtkinter.CTkFrame(main_content_frame,width=Main_app.winfo_screenwidth(),height=Main_app.winfo_screenheight(),fg_color="#F2E8BD")
library_frame = customtkinter.CTkFrame(main_content_frame,width=Main_app.winfo_screenwidth(),height=Main_app.winfo_screenheight(),fg_color="#F2E8BD")
add_books_frame = customtkinter.CTkFrame(main_content_frame,width=Main_app.winfo_screenwidth(),height=Main_app.winfo_screenheight(),fg_color="#F2E8BD")

# Side bar Frames
title_frame = customtkinter.CTkFrame(Main_app,width=150,height=Main_app.winfo_screenheight(),fg_color="#C8DF8C",corner_radius=0,border_width=1,border_color="#C8DF8C")
title_frame.place(x = -75, y = 0)

icon_frame = customtkinter.CTkFrame(Main_app,width=76,height=Main_app.winfo_screenheight(),fg_color="#C8DF8C",corner_radius=0,border_width=1,border_color="#C8DF8C")
icon_frame.place(x = 0, y = 0)

# Credential Frames
sign_in_frame = customtkinter.CTkFrame(Main_app,width=Main_app.winfo_screenwidth(),height=Main_app.winfo_screenheight(),fg_color="#F2E8BD")

log_in_frame = customtkinter.CTkFrame(Main_app,width=Main_app.winfo_screenwidth(),height=Main_app.winfo_screenheight(),fg_color="#F2E8BD")
log_in_frame.pack(fill='both', expand=1)    

# Sign in frame
sign_in_label_img = customtkinter.CTkImage(Image.open('D:\Code\DSA\Final-Project-DSA-main\STACKRead\icons\stackreadicon.png'),size=(75,75))
sign_in_label = customtkinter.CTkLabel(sign_in_frame,text="",image=sign_in_label_img,fg_color="transparent")
sign_in_label.place(x=390,y=100)

sign_in_icon_stack_label_img = customtkinter.CTkImage(Image.open('D:\Code\DSA\Final-Project-DSA-main\STACKRead\icons\logo.png'),size=(475,75))
sign_in_icon_stack_label = customtkinter.CTkLabel(sign_in_frame,image=sign_in_icon_stack_label_img,anchor="w",text="",fg_color="transparent",font=("Quando",25))
sign_in_icon_stack_label.place(x=470,y=100)

sign_in_icon_quote_label = customtkinter.CTkLabel(sign_in_frame,anchor="w",text="Your digital book wiki.",fg_color="transparent",font=("Quando",25))
sign_in_icon_quote_label.place(x=640,y=180)

sign_in_credentials_frame = customtkinter.CTkFrame(sign_in_frame,fg_color="#1E1E1E",corner_radius=25,width=725,height=425)
sign_in_credentials_frame.place(x=310,y=220)

sign_in_acc_reg_label = customtkinter.CTkLabel(sign_in_credentials_frame,anchor="center",text_color="white",text="Account Registration",fg_color="transparent",font=("Quando",45))
sign_in_acc_reg_label.place(x=110,y=30)

sign_in_username_entry = customtkinter.CTkEntry(sign_in_credentials_frame,fg_color="white",placeholder_text="Username",font=("Quando",25),corner_radius=25,width=630,height=50)
sign_in_username_entry.place(x=50,y=100)

sign_in_email_entry = customtkinter.CTkEntry(sign_in_credentials_frame,width=630,height=50,fg_color="white",placeholder_text="Email",corner_radius=25,font=("Quando",25))
sign_in_email_entry.place(x=50,y=160)

sign_in_password_entry = customtkinter.CTkEntry(sign_in_credentials_frame,width=630,height=50,fg_color="white",placeholder_text="Password",corner_radius=25,show="●",font=("Quando",25))
sign_in_password_entry.place(x=50,y=220)

terms_n_conditions = False
sign_in_checkbox = customtkinter.CTkCheckBox(sign_in_credentials_frame,command=agree, onvalue='agree',offvalue='disagree',text_color="white",width=170,height=50,text="I have read and understood the terms and conditions",font=("Quando",15))
sign_in_checkbox.place(x=150,y=270)

sign_register_in_button = customtkinter.CTkButton(sign_in_credentials_frame,command=lambda: sign_in(),text="Register",width=230,height=50,fg_color="#C8DF8C",corner_radius=25,text_color="black",font=("Quando",25))
sign_register_in_button.place(x=250,y=340)

sign_in_already_have_an_account_label = customtkinter.CTkLabel(sign_in_frame,anchor="w",text_color="black",text="Already have an account?",fg_color="transparent",font=("Quando",20))
sign_in_already_have_an_account_label.place(x=410,y=655)

clock_in_button = customtkinter.CTkButton(sign_in_frame,command=lambda: to_log_in(),text="Clock-In",width=150,height=25,fg_color="#606060",corner_radius=25,text_color="white",font=("Quando",20))
clock_in_button.place(x=700,y=655)

# Log in Frame Content
login_icon_label_img = customtkinter.CTkImage(Image.open('D:\Code\DSA\Final-Project-DSA-main\STACKRead\icons\stackreadicon.png'),size=(75,75))
login_icon_label = customtkinter.CTkLabel(log_in_frame,text="",image=login_icon_label_img,fg_color="transparent")
login_icon_label.place(x=390,y=100)

login_icon_stack_label_img = customtkinter.CTkImage(Image.open('D:\Code\DSA\Final-Project-DSA-main\STACKRead\icons\logo.png'),size=(475,75))
login_icon_stack_label = customtkinter.CTkLabel(log_in_frame,image=login_icon_stack_label_img,anchor="w",text="",fg_color="transparent",font=("Quando",25))
login_icon_stack_label.place(x=470,y=100)

login_icon_quote_label = customtkinter.CTkLabel(log_in_frame,anchor="w",text="Your digital book wiki.",fg_color="transparent",font=("Quando",25))
login_icon_quote_label.place(x=640,y=180)

login_credentials_frame = customtkinter.CTkFrame(log_in_frame,fg_color="#B88B68",corner_radius=25,width=725,height=525)
login_credentials_frame.place(x=310,y=220)

login_username_entry = customtkinter.CTkEntry(login_credentials_frame,fg_color="white",placeholder_text="Username",font=("Quando",25),corner_radius=25,width=630,height=50)
login_username_entry.place(x=50,y=40)

login_password_entry = customtkinter.CTkEntry(login_credentials_frame,width=630,height=50,fg_color="white",placeholder_text="Password",corner_radius=25,show="●",font=("Quando",25))
login_password_entry.place(x=50,y=100)

login_button = customtkinter.CTkButton(login_credentials_frame,command=lambda: login(),text="Clock-in",width=230,height=50,fg_color="#C8DF8C",corner_radius=25,text_color="black",font=("Quando",25))
login_button.place(x=250,y=170)

login_clock_frame = customtkinter.CTkFrame(log_in_frame,width=825,height=325,fg_color="#D9D9D9",corner_radius=25)
login_clock_frame.place(x=260,y=470)

login_clock_label = customtkinter.CTkLabel(login_clock_frame,text="12:00 AM",width=725,height=195,text_color="white",fg_color="#292D32",corner_radius=25,font=("Quando",65))
login_clock_label.place(x=50,y=50)

login_no_account_label = customtkinter.CTkLabel(login_clock_frame,text="Don't have an account?",width=150,height=25,fg_color="transparent",text_color="black",corner_radius=25,font=("Quando",20))
login_no_account_label.place(x=140,y=255)

login_register_button = customtkinter.CTkButton(login_clock_frame,text="Register",command=to_sign_in,corner_radius=450,width=150,height=25,fg_color="#9A9A9A",text_color="black",font=("Quando",20))
login_register_button.place(x=440,y=255)
log_in_clock()

# Main Content Frames

# Home Content
searchbarx=((window_width/2)-325)
home_search_entry_frame = customtkinter.CTkFrame(home_frame,width=650,height=50,fg_color="transparent",corner_radius=45)
home_search_entry_frame.place(x = searchbarx, y = 20)

home_search_entry = customtkinter.CTkEntry(home_search_entry_frame,width=550,height=50,fg_color="white",corner_radius=45,placeholder_text="Search Titles...",font=("Quando",20))
home_search_entry.place(x = 0,y=0)

home_search_entry_button_img = customtkinter.CTkImage((Image.open("D:\Code\DSA\Final-Project-DSA-main\STACKRead\icons\\searchicon.png")),size=(30,30))
home_search_entry_button = customtkinter.CTkButton(home_search_entry_frame,text="",image=home_search_entry_button_img,width=30,height=30,fg_color="white",corner_radius=0,border_color="white",hover_color="white")
home_search_entry_button.place(x = 500 , y=5)

home_bookmark_frame = customtkinter.CTkFrame(home_frame, width=453, height=120, fg_color="#606060", corner_radius=0)
home_bookmark_frame.place(x = 570, y = 80)

home_user_frame = customtkinter.CTkFrame(home_frame, width=310, height=120, fg_color="#B88B68", corner_radius=0)
home_user_frame.place(x = 358, y = 80)

home_user_label = customtkinter.CTkLabel(home_user_frame,anchor=customtkinter.W,font=("Quando",25),text_color="white",width=100,height=30,fg_color="transparent",corner_radius=0, text="my Username")
home_user_label.place(x = 75, y = 30)

home_date_label = customtkinter.CTkLabel(home_user_frame,anchor=customtkinter.W,font=("Quando",15),text_color="white",width=100,height=10,fg_color="transparent",corner_radius=0, text="Date Joined: " )
home_date_label.place(x = 75, y = 70)

home_bookmark_label = customtkinter.CTkLabel(home_bookmark_frame,anchor=customtkinter.W,font=("Quando",20),text_color="white",width=100,height=30,fg_color="transparent",corner_radius=0,text="Bookmark: " )
home_bookmark_label.place(x = 136, y = 40)

home_upload_label = customtkinter.CTkLabel(home_bookmark_frame,anchor=customtkinter.W,font=("Quando",20),text_color="white",width=100,height=30,fg_color="transparent",corner_radius=0,text="Uploaded: " )
home_upload_label.place(x = 286, y = 40)

home_upload_frame_label_img = customtkinter.CTkImage((Image.open("D:\Code\DSA\Final-Project-DSA-main\STACKRead\icons\\camera.png")), size=(25,25))
home_upload_frame_img = customtkinter.CTkButton(home_frame, text="", image=home_upload_frame_label_img, width=120,height=120, fg_color="#614D40", corner_radius=0)
home_upload_frame_img.place(x = 270, y = 80)

home_lithit_frame_outer = customtkinter.CTkFrame(home_frame,width=485,height=510,fg_color="transparent",corner_radius=0)
home_lithit_frame_outer.place(x = 176, y = 260)

home_lithit_frame_inner = customtkinter.CTkFrame(home_lithit_frame_outer,width=750,height=642,fg_color="#B88B68",corner_radius=0)
home_lithit_frame_inner.place(x = 10, y = 40)

home_lithit_firstshelf = customtkinter.CTkFrame(home_lithit_frame_outer, width=420, height=100, fg_color="#614D40", corner_radius=0)
home_lithit_firstshelf.place(x = 35, y = 60)

home_lithit_secondshelf = customtkinter.CTkFrame(home_lithit_frame_outer, width=420, height=100, fg_color="#614D40", corner_radius=0)
home_lithit_secondshelf.place(x = 35, y = 170)

home_lithit_thirdshelf = customtkinter.CTkFrame(home_lithit_frame_outer, width=420, height=100, fg_color="#614D40", corner_radius=0)
home_lithit_thirdshelf.place(x = 35, y = 280)

home_lithit_fourthshelf = customtkinter.CTkFrame(home_lithit_frame_outer, width=420, height=100, fg_color="#614D40", corner_radius=0)
home_lithit_fourthshelf.place(x = 35, y = 390)

home_lithit_button = customtkinter.CTkButton(home_lithit_frame_outer,width=150,height=30,font=("Quando",20),fg_color="#606060",hover_color="#4c4c4c",corner_radius=0,text="Literary Hits")
home_lithit_button.place(x = 10, y = 10)

home_arrowlit_button_img = customtkinter.CTkImage((Image.open("D:\Code\DSA\Final-Project-DSA-main\STACKRead\icons\\kanan.png")))
home_arrowlit_button_label = customtkinter.CTkButton(home_frame, image=home_arrowlit_button_img, width=25, height=25, fg_color="transparent",hover_color="#F2E8BD", corner_radius=0, text="")
home_arrowlit_button_label.place(x = 635, y = 270)

home_lastread_frame_outer = customtkinter.CTkFrame(home_frame,width=485,height=510,fg_color="transparent",corner_radius=0)
home_lastread_frame_outer.place(x = 670, y = 260)

home_lastread_frame_inner = customtkinter.CTkFrame(home_lastread_frame_outer,width=642,height=642,fg_color="#B88B68",corner_radius=0)
home_lastread_frame_inner.place(x = 10, y = 40)

home_lastread_firstshelf = customtkinter.CTkFrame(home_lastread_frame_outer, width=420, height=100, fg_color="#614D40", corner_radius=0)
home_lastread_firstshelf.place(x = 35, y = 60)

home_lastread_secondshelf = customtkinter.CTkFrame(home_lastread_frame_outer, width=420, height=100, fg_color="#614D40", corner_radius=0)
home_lastread_secondshelf.place(x = 35, y = 170)

home_lastread_thirdshelf = customtkinter.CTkFrame(home_lastread_frame_outer, width=420, height=100, fg_color="#614D40", corner_radius=0)
home_lastread_thirdshelf.place(x = 35, y = 280)

home_lastread_fourthshelf = customtkinter.CTkFrame(home_lastread_frame_outer, width=420, height=100, fg_color="#614D40", corner_radius=0)
home_lastread_fourthshelf.place(x = 35, y = 390)

home_lastread_button = customtkinter.CTkButton(home_lastread_frame_outer,width=150,height=30,font=("Quando",20),fg_color="#606060",hover_color="#4c4c4c",corner_radius=0,text="Last Read")
home_lastread_button.place(x = 10, y = 10)

home_arrowlast_button_img = customtkinter.CTkImage((Image.open("D:\Code\DSA\Final-Project-DSA-main\STACKRead\icons\\kanan.png")))
home_arrowlast_button_label = customtkinter.CTkButton(home_frame, image=home_arrowlast_button_img, width=25, height=25, fg_color="transparent",hover_color="#F2E8BD", corner_radius=0, text="")
home_arrowlast_button_label.place(x = 1130, y = 270)

# lithit Content
lithit_main_frame = customtkinter.CTkFrame(lithit_frame, width=1360, height=765,fg_color="#F2E8BD")
lithit_main_frame.place(x=0, y=0)

lithit_search_entry_frame = customtkinter.CTkFrame(lithit_main_frame,width=650,height=50,fg_color="transparent",corner_radius=45)
lithit_search_entry_frame.place(x = 420, y = 20)

lithit_search_entry = customtkinter.CTkEntry(lithit_search_entry_frame,width=550,height=50,fg_color="white",corner_radius=45,placeholder_text="Search Titles...",font=("Quando",20))
lithit_search_entry.place(x = 0,y=0)

lithit_search_entry_button_img = customtkinter.CTkImage((Image.open("D:\Code\DSA\Final-Project-DSA-main\STACKRead\icons\searchicon.png")),size=(30,30))
lithit_search_entry_button = customtkinter.CTkButton(lithit_search_entry,text="",image=lithit_search_entry_button_img,width=30,height=30,fg_color="white",corner_radius=0,border_color="white",hover_color="white")
lithit_search_entry_button.place(x = 500 , y=5)

lithit_firstShelf_frame = customtkinter.CTkScrollableFrame(lithit_main_frame, width=1000 , height=635, fg_color="#B88B68")
lithit_firstShelf_frame.place(x=170, y=120)

update_last_ten_stack()

# Search Content
search_advanceSearch_frame = customtkinter.CTkFrame(search_frame,width=700,height=1080,fg_color="#433F3D",corner_radius=0)
search_advanceSearch_frame.place(x = (window_width/2)-710, y = 0)

search_library_frame_inner = customtkinter.CTkFrame(search_frame,width=1509,height=1122,fg_color="#B88B68",corner_radius=0)
search_library_frame_inner.place(x = 600, y = 0)
        
search_advanceSearch_label = customtkinter.CTkLabel(search_advanceSearch_frame, anchor=customtkinter.W, font=("Quando", 25), text_color="white", width=450, height=69, fg_color="transparent", corner_radius=0, text="Advance Search")
search_advanceSearch_label.place(x=180, y=20)

search_genre_label = customtkinter.CTkLabel(search_advanceSearch_frame, anchor=customtkinter.E, font=("Quando", 20), text_color="white", width=70, height=30, fg_color="transparent", corner_radius=0, text="Genre")
search_genre_label.place(x=225, y=100)
        
search_genre_entry = customtkinter.CTkEntry(search_advanceSearch_frame,width=200,height=30,fg_color="white",corner_radius=0,placeholder_text="",font=("Quando",20),border_width=0)
search_genre_entry.place(x = 300,y = 100)

search_author_label = customtkinter.CTkLabel(search_advanceSearch_frame, anchor=customtkinter.E, font=("Quando", 20), text_color="white", width=70, height=30, fg_color="transparent", corner_radius=0, text="Author")
search_author_label.place(x=225, y=150)
        
search_author_entry = customtkinter.CTkEntry(search_advanceSearch_frame,width=200,height=30,fg_color="white",corner_radius=0,placeholder_text="",font=("Quando",20),border_width=0)
search_author_entry.place(x = 300,y = 150)

search_theme_label = customtkinter.CTkLabel(search_advanceSearch_frame, anchor=customtkinter.E, font=("Quando", .00002 * (window_width * window_height)), text_color="white", width=70, height=30, fg_color="transparent", corner_radius=0, text="Theme")
search_theme_label.place(x=225, y=200)
        
search_theme_entry = customtkinter.CTkEntry(search_advanceSearch_frame,width=200,height=30,fg_color="white",corner_radius=0,placeholder_text="",font=("Quando",20),border_width=0)
search_theme_entry.place(x = 300,y = 200)
        
search_rating_label = customtkinter.CTkLabel(search_advanceSearch_frame, anchor=customtkinter.E, font=("Quando", 20), text_color="white", width=70, height=30, fg_color="transparent", corner_radius=0, text="Rating")
search_rating_label.place(x=225, y=250)
        
search_rating_entry = customtkinter.CTkEntry(search_advanceSearch_frame,width=200,height=30,fg_color="white",corner_radius=0,placeholder_text="",font=("Quando",20),border_width=0)
search_rating_entry.place(x = 300,y = 250)

search_publishDate_label = customtkinter.CTkLabel(search_advanceSearch_frame, anchor=customtkinter.E, font=("Quando",20), text_color="white", width=70, height=30, fg_color="transparent", corner_radius=0, text="Publish Date")
search_publishDate_label.place(x=185, y=300)
        
search_publishDate_entry = customtkinter.CTkEntry(search_advanceSearch_frame,width=200,height=30,fg_color="white",corner_radius=0,placeholder_text="",font=("Quando",20),border_width=0)
search_publishDate_entry.place(x = 300,y = 300)

search_resetFilter_button = customtkinter.CTkButton(search_advanceSearch_frame,width=150,height=40,font=("Quando",20),fg_color="#c0c0c0",text_color="black" ,hover_color="dark grey" ,corner_radius=60,text="Reset")
search_resetFilter_button.place(x = 150, y = 350)   

search_search_button = customtkinter.CTkButton(search_advanceSearch_frame,width=150,height=40,font=("Quando",20),fg_color="#C8DF8C",text_color="black" ,hover_color="#8c9c62" ,corner_radius=60,text="Search")
search_search_button.place(x = 330, y = 350)

# Library Content

library_main_frame = customtkinter.CTkFrame(library_frame,width=1360,height=765,fg_color="#F2E8BD",corner_radius=0)
library_main_frame.place(x = 0, y = 0)

library_library_inner_frame = customtkinter.CTkFrame(library_main_frame,width=1000,height=800,fg_color="#433F3D",corner_radius=0)
library_library_inner_frame.place(x = 170, y = 70)

library_bookshelf_frame = customtkinter.CTkScrollableFrame(library_library_inner_frame,width=900,height=610,fg_color="#B88B68")
library_bookshelf_frame.place(x = 40, y = 78)

display_stack_as_buttons()

library_search_entry_frame = customtkinter.CTkFrame(library_main_frame,width=650,height=50,fg_color="transparent",corner_radius=45)
library_search_entry_frame.place(x = 450, y = 20)

library_search_entry = customtkinter.CTkEntry(library_search_entry_frame,width=450,height=45,fg_color="white",corner_radius=45,placeholder_text="Search Titles...",font=("Quando",20))
library_search_entry.place(x = 0,y=0)

library_search_entry_button_img = customtkinter.CTkImage((Image.open("D:\Code\DSA\Final-Project-DSA-main\STACKRead\icons\\searchicon.png")),size=(55,55))
library_search_entry_button = customtkinter.CTkButton(library_search_entry,text="",image=home_search_entry_button_img,width=10,height=10,fg_color="transparent",hover_color="white")
library_search_entry_button.place(x = 390 , y=2.5)

library_reading_button = customtkinter.CTkButton(library_library_inner_frame,width=100,height=40,font=("Quando",20),fg_color="#433F3D", hover_color="#B88B68",corner_radius=0,text="Reading")
library_reading_button.place(x = 5, y = 5)

library_plan2read_button = customtkinter.CTkButton(library_library_inner_frame,width=140,height=40,font=("Quando",20),fg_color="#433F3D", hover_color="#B88B68",corner_radius=0,text="Plan to read")
library_plan2read_button.place(x = 100, y = 5)

library_completed_button = customtkinter.CTkButton(library_library_inner_frame,width=120,height=40,font=("Quando",20),fg_color="#433F3D", hover_color="#B88B68",corner_radius=0,text="Completed")
library_completed_button.place(x = 240, y = 5)

library_dropped_button = customtkinter.CTkButton(library_library_inner_frame,width=120,height=40,font=("Quando",20),fg_color="#433F3D", hover_color="#B88B68",corner_radius=0,text="Dropped")
library_dropped_button.place(x = 360, y = 5)   

# Addbooks Content
add_books_main_frame = customtkinter.CTkFrame(add_books_frame,width=1260,height=760,fg_color="transparent",corner_radius=0)
add_books_main_frame.place(x=60,y=0)

addbooks_upload_cover_frame = customtkinter.CTkFrame(add_books_main_frame,width=250,height=360,fg_color="#414240",corner_radius=0)
addbooks_upload_cover_frame.place(x=140,y=10)

addbooks_upload_cover_button_img = customtkinter.CTkImage(Image.open('D:\Code\DSA\Final-Project-DSA-main\STACKRead\icons\coverbookupload.png'),size=(25,25))
addbooks_upload_cover_button = customtkinter.CTkButton(add_books_main_frame,image=addbooks_upload_cover_button_img,text_color="black",font=("Quando",15),text="Upload Cover",width=250,height=50,fg_color="#C8C8C8",corner_radius=25)
addbooks_upload_cover_button.place(x=140,y=380)

addbooks_book_detail_input_frame = customtkinter.CTkFrame(add_books_main_frame,width=715,height=420,fg_color="#808080",corner_radius=0)
addbooks_book_detail_input_frame.place(x=405,y=10)

addbooks_book_detail_input_title_entry = customtkinter.CTkEntry(addbooks_book_detail_input_frame,font=("Quando",25),width=695,height=55,fg_color="#F1F1F1",corner_radius=25,placeholder_text="Input Book Title",placeholder_text_color="gray")
addbooks_book_detail_input_title_entry.place(x=10,y=10)

addbooks_book_detail_input_author_name_entry = customtkinter.CTkEntry(addbooks_book_detail_input_frame,font=("Quando",15),width=525,height=45,fg_color="#F1F1F1",corner_radius=25,placeholder_text="Input Author(s) Name",placeholder_text_color="gray")
addbooks_book_detail_input_author_name_entry.place(x=10,y=75)

addbooks_book_detail_input_publish_date_entry = customtkinter.CTkEntry(addbooks_book_detail_input_frame,font=("Quando",12),width=165,height=45,fg_color="#F1F1F1",corner_radius=25,placeholder_text="Publish Date",placeholder_text_color="gray")
addbooks_book_detail_input_publish_date_entry.place(x=540,y=75)

addbooks_book_detail_input_inner_frame = customtkinter.CTkFrame(addbooks_book_detail_input_frame,width=715,height=290,fg_color="#C8C8C8",corner_radius=0)
addbooks_book_detail_input_inner_frame.place(x=0,y=130)

addbooks_book_detail_input_genre_label = customtkinter.CTkLabel(addbooks_book_detail_input_inner_frame,text="Genre:",font=("Quando",20),width=75,height=55,fg_color="transparent")
addbooks_book_detail_input_genre_label.place(x=10,y=10)

addbooks_book_detail_input_theme_label = customtkinter.CTkLabel(addbooks_book_detail_input_inner_frame,text="Theme:",font=("Quando",20),width=75,height=55,fg_color="transparent")
addbooks_book_detail_input_theme_label.place(x=10,y=65)

addbooks_book_detail_input_keywords_label = customtkinter.CTkLabel(addbooks_book_detail_input_inner_frame,text="Keywords:",font=("Quando",20),width=75,height=55,fg_color="transparent")
addbooks_book_detail_input_keywords_label.place(x=10,y=120)

addbooks_book_detail_input_language_label = customtkinter.CTkLabel(addbooks_book_detail_input_inner_frame,text="Language:",font=("Quando",20),width=75,height=55,fg_color="transparent")
addbooks_book_detail_input_language_label.place(x=10,y=175)

addbooks_book_detail_input_rating_label = customtkinter.CTkLabel(addbooks_book_detail_input_inner_frame,text="Rating:",font=("Quando",20),width=75,height=55,fg_color="transparent")
addbooks_book_detail_input_rating_label.place(x=10,y=235)

addbooks_book_detail_input_genre_entry = customtkinter.CTkEntry(addbooks_book_detail_input_inner_frame,font=("Quando",12),width=565,height=45,fg_color="#F1F1F1",corner_radius=25)
addbooks_book_detail_input_genre_entry.place(x=135,y=10)

addbooks_book_detail_input_theme_entry = customtkinter.CTkEntry(addbooks_book_detail_input_inner_frame,font=("Quando",12),width=565,height=45,fg_color="#F1F1F1",corner_radius=25)
addbooks_book_detail_input_theme_entry.place(x=135,y=65)

addbooks_book_detail_input_keywords_entry = customtkinter.CTkEntry(addbooks_book_detail_input_inner_frame,font=("Quando",12),width=565,height=45,fg_color="#F1F1F1",corner_radius=25)
addbooks_book_detail_input_keywords_entry.place(x=135,y=120)

addbooks_book_detail_input_rating_entry = customtkinter.CTkEntry(addbooks_book_detail_input_inner_frame,font=("Quando",12),width=565,height=45,fg_color="#F1F1F1",corner_radius=25)
addbooks_book_detail_input_rating_entry.place(x=135,y=175)

addbooks_book_detail_rating = customtkinter.StringVar(value="other")

my_rad1 = customtkinter.CTkRadioButton(addbooks_book_detail_input_inner_frame, text="1", value="1", variable=addbooks_book_detail_rating,)
my_rad1.place(x=155,y=245)

my_rad2 = customtkinter.CTkRadioButton(addbooks_book_detail_input_inner_frame, text="2", value="2", variable=addbooks_book_detail_rating,)
my_rad2.place(x=215,y=245)

my_rad3 = customtkinter.CTkRadioButton(addbooks_book_detail_input_inner_frame, text="3", value="3", variable=addbooks_book_detail_rating,)
my_rad3.place(x=265,y=245)

my_rad4 = customtkinter.CTkRadioButton(addbooks_book_detail_input_inner_frame, text="4", value="4", variable=addbooks_book_detail_rating,)
my_rad4.place(x=315,y=245)

my_rad5 = customtkinter.CTkRadioButton(addbooks_book_detail_input_inner_frame, text="5", value="5", variable=addbooks_book_detail_rating,)
my_rad5.place(x=365,y=245)

addbooks_synopsis_frame = customtkinter.CTkFrame(add_books_main_frame,width=980,height=450,fg_color="#808080",corner_radius=0)
addbooks_synopsis_frame.place(x=140,y=445)

addbooks_synopsis_frame_label = customtkinter.CTkLabel(addbooks_synopsis_frame,anchor="w",text="Synopsis",width=980,height=75,fg_color="transparent",corner_radius=0,text_color="white",font=("Quando",35))
addbooks_synopsis_frame_label.place(x=20,y=0)

addbooks_synopsis_textbox = customtkinter.CTkTextbox(addbooks_synopsis_frame,width=980,height=235,fg_color="#D9D9D9",corner_radius=0,font=("Quando",25))
addbooks_synopsis_textbox.place(x=0,y=85)

addbooks_upload_button_img = customtkinter.CTkImage((Image.open("D:\Code\DSA\Final-Project-DSA-main\STACKRead\icons\\uploadpic.png")),size=(35,35))
addbooks_upload_button = customtkinter.CTkButton(add_books_main_frame,image=addbooks_upload_button_img,fg_color="#B88B68",text="",width=50,height=75,corner_radius=0,hover_color="#737373")
addbooks_upload_button.place(x =1120, y = 685)

# Title and Icon Frame Content
button_stackread_img = customtkinter.CTkImage((Image.open("D:\Code\DSA\Final-Project-DSA-main\STACKRead\icons\\stackreadicon.png")),size=(45,45))
button_stackread = customtkinter.CTkButton(icon_frame,image=button_stackread_img,text="",width=75,height=75,command=side_panel_switch,corner_radius=0,fg_color="#C8DF8C",hover_color="#737373",hover=False)
button_stackread.place(x = 0, y = 10)

label_title_img = customtkinter.CTkImage(Image.open('D:\Code\DSA\Final-Project-DSA-main\STACKRead\icons\\logo.png'),size=(145,25))
label_title = customtkinter.CTkLabel(title_frame,image=label_title_img,anchor=customtkinter.W, text="",width=150,height=75,font=("Quando",20),fg_color="transparent",text_color="black")
label_title.place(x = 0, y = 10)

# Buttons for sidebar
button_home_img = customtkinter.CTkImage((Image.open("D:\Code\DSA\Final-Project-DSA-main\STACKRead\icons\\homeunclicked.png")),size=(45,45))
button_home = customtkinter.CTkButton(icon_frame, image=button_home_img,text="", command=lambda: show_main_frame_content("Home"),width=75,height=75,corner_radius=0,fg_color="#C8DF8C",hover=False)
button_home.place(x = 0 , y = 100)

button_lithit_img = customtkinter.CTkImage((Image.open("D:\Code\DSA\Final-Project-DSA-main\STACKRead\icons\\lithitsunclicked.png")),size=(35,45))
button_lithit = customtkinter.CTkButton(icon_frame,image=button_lithit_img, text="", command=lambda: show_main_frame_content("lithit"),width=75,height=75,corner_radius=0,font=("Quando",16),fg_color="#C8DF8C",hover=False)
button_lithit.place(x = 0 , y = 180)

button_search_img = customtkinter.CTkImage((Image.open("D:\Code\DSA\Final-Project-DSA-main\STACKRead\icons\\searchunclicked.png")),size=(45,45))
button_search = customtkinter.CTkButton(icon_frame,image=button_search_img, text="", command=lambda: show_main_frame_content("Search"),width=75,height=75,corner_radius=0,font=("Quando",16),fg_color="#C8DF8C",hover=False)
button_search.place(x = 0 , y = 260)

button_library_img = customtkinter.CTkImage((Image.open("D:\Code\DSA\Final-Project-DSA-main\STACKRead\icons\\libraryunclicked.png")),size=(45,45))
button_library = customtkinter.CTkButton(icon_frame,image=button_library_img, text="", command=lambda: show_main_frame_content("Library"),width=75,height=75,corner_radius=0,font=("Quando",16),fg_color="#C8DF8C",hover=False)
button_library.place(x = 0 , y = 340)

button_addbooks_img= customtkinter.CTkImage((Image.open("D:\Code\DSA\Final-Project-DSA-main\STACKRead\icons\\disclickedaddbooks.png")),size=(45,45))
button_addbooks = customtkinter.CTkButton(icon_frame,image=button_addbooks_img,text="",command=lambda: show_main_frame_content("Add"),width=75,height=75,corner_radius=0,font=("Quando",16),fg_color="#C8DF8C",hover=False)
button_addbooks.place(x = 0, y = 420)

button_log_out_img= customtkinter.CTkImage((Image.open("D:\Code\DSA\Final-Project-DSA-main\STACKRead\icons\\logout.png")),size=(35,35))
button_log_out = customtkinter.CTkButton(icon_frame,image=button_log_out_img, text="", command=lambda: logout(),width=75,height=75,corner_radius=0,font=("Quando",16),fg_color="#C8DF8C",hover=False,text_color="black")
button_log_out.place(x = 5, y = 675)

# Button Labels for Sidebar
button_label_home = customtkinter.CTkButton(title_frame,anchor="w", text="Home",width=150,height=75,corner_radius=0,fg_color="transparent",font=("Quando",15),text_color="black",hover=False)
button_label_home.place(x = 0 , y = 100)

button_label_lithit = customtkinter.CTkButton(title_frame,anchor="w", text="Literary hits",width=150,height=75,corner_radius=0,fg_color="transparent",font=("Quando",15),text_color="black",hover=False)
button_label_lithit.place(x = 0 , y = 180)

button_label_search = customtkinter.CTkButton(title_frame,anchor="w", text="Browse",width=150,height=75,corner_radius=0,fg_color="transparent",font=("Quando",15),text_color="black",hover=False)
button_label_search.place(x = 0 , y = 260)

button_label_library = customtkinter.CTkButton(title_frame,anchor="w", text="Library",width=150,height=75,corner_radius=0,fg_color="transparent",font=("Quando",15),text_color="black",hover=False)
button_label_library.place(x = 0 , y = 340)

button_label_addbooks = customtkinter.CTkButton(title_frame,anchor="w", text="Add books",width=150,height=75,corner_radius=0,fg_color="transparent",font=("Quando",15),text_color="black",hover=False)
button_label_addbooks.place(x = 0, y = 420)

# button_label_log_out_img= customtkinter.CTkImage((Image.open("icons\disclickedaddbooks.png")),size=(50,50))
button_label_log_out = customtkinter.CTkButton(title_frame,anchor="w", text="Log out",width=150,height=75,corner_radius=0,fg_color="transparent",font=("Quando",15),text_color="black",hover=False)
button_label_log_out.place(x = 0, y = 675)

# Main Loop
Main_app.mainloop()
